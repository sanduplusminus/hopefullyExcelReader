import openpyxl

workbook = openpyxl.load_workbook(r"D:\clasa9aNOTE.xlsx")

sheet = workbook.active

command_main = input(str("command: "))
command_lesson = input(str("lesson: "))
command_period = input(str("period: "))
command_showGrades = input(str("show grades?: "))
command_showGrades = command_showGrades.strip().lower()[0] == "y"

print("")


def FIND_LESSON_AVG(lesson, period, showGrades):
   # get the column number for the lesson
   grade_column = search_for_lesson(lesson)
   if grade_column is None:
       return(f"lesson '{lesson}' not found.")

   # get the starting row based on the period
   overall_is_true = False
   semester_wait = "none"

   if period.strip().lower()[0] == "m":
       starting_row = search_for_starting_row(period, grade_column)
       if starting_row is None:
            return(f"period '{period}' not found for lesson '{lesson}'.")
   elif period.strip().lower()[0] == "s":
       if period.strip().lower() == "s1":
           starting_row = starting_row = search_for_starting_row("m1", grade_column)
           semester_wait = "m5"
       elif period.strip().lower() == "s2":
           starting_row = search_for_starting_row("m5", grade_column)
           semester_wait = "placeholder"
       else:
           return(f"period '{period}' not found for lesson '{lesson}'.")
   else:
       starting_row = 3
       overall_is_true = True
   
   total = 0
   count = 0

   # start from row 3
   for row in range(starting_row + 1, sheet.max_row + 1):
       cell_value = sheet.cell(row=row, column=grade_column).value
       if semester_wait == "none":
           if cell_value is not None:
               if isinstance(cell_value, (int, float)):
                   total += cell_value
                   count += 1
                   if showGrades:
                       print(cell_value)
               elif overall_is_true == False:
                   break
       else: # calculate until the semester is reached
           if cell_value is not None:
               if isinstance(cell_value, (int, float)):
                   total += cell_value
                   count += 1
                   if showGrades:
                       print(cell_value)
               elif cell_value.strip().lower() == semester_wait.strip().lower():
                   break


   if count > 0:
       average = total / count
       return(f"average for '{lesson}' in period '{period}': {average:.2f}")
   else:
       return("no valid grades found.")



def FIND_LESSON_AVG_AS_FLOAT(lesson, period, showGrades):
   # get the column number for the lesson
   grade_column = search_for_lesson(lesson)
   if grade_column is None:
       return(f"lesson '{lesson}' not found.")

   # get the starting row based on the period
   overall_is_true = False
   semester_wait = "none"

   if period.strip().lower()[0] == "m":
       starting_row = search_for_starting_row(period, grade_column)
       if starting_row is None:
            return 0
   elif period.strip().lower()[0] == "s":
       if period.strip().lower() == "s1":
           starting_row = starting_row = search_for_starting_row("m1", grade_column)
           semester_wait = "m5"
       elif period.strip().lower() == "s2":
           starting_row = search_for_starting_row("m5", grade_column)
           semester_wait = "amogus"
       else:
           return 0
   else:
       starting_row = 3
       overall_is_true = True
   
   total = 0
   count = 0

   # start from row 3
   for row in range(starting_row + 1, sheet.max_row + 1):
       cell_value = sheet.cell(row=row, column=grade_column).value
       if semester_wait == "none":
           if cell_value is not None:
               if isinstance(cell_value, (int, float)):
                   total += cell_value
                   count += 1
               elif overall_is_true == False:
                   break
       else: # calculate until the semester is reached
           if cell_value is not None:
               if isinstance(cell_value, (int, float)):
                   total += cell_value
                   count += 1
               elif cell_value.strip().lower() == semester_wait.strip().lower():
                   break


   if count > 0:
       average = total / count
       return(round(average, 2))
   else:
       return 0



def FIND_LESSON_COUNT(lesson, period, showGrades):
   # get the column number for the lesson
   grade_column = search_for_lesson(lesson)
   if grade_column is None:
       return(f"lesson '{lesson}' not found.")

   # get the starting row based on the period
   overall_is_true = False
   semester_wait = "none"

   if period.strip().lower()[0] == "m":
       starting_row = search_for_starting_row(period, grade_column)
       if starting_row is None:
            return(f"period '{period}' not found for lesson '{lesson}'.")
   elif period.strip().lower()[0] == "s":
       if period.strip().lower() == "s1":
           starting_row = starting_row = search_for_starting_row("m1", grade_column)
           semester_wait = "m5"
       elif period.strip().lower() == "s2":
           starting_row = search_for_starting_row("m5", grade_column)
           semester_wait = "placeholder"
       else:
           return(f"period '{period}' not found for lesson '{lesson}'.")
   else:
       starting_row = 3
       overall_is_true = True
   
   count = 0

   # start from row 3
   for row in range(starting_row + 1, sheet.max_row + 1):
       cell_value = sheet.cell(row=row, column=grade_column).value
       if semester_wait == "none":
           if cell_value is not None:
               if isinstance(cell_value, (int, float)):
                   count += 1
                   if showGrades:
                       print(cell_value)
               elif overall_is_true == False:
                   break
       else: # calculate until the semester is reached
           if cell_value is not None:
               if isinstance(cell_value, (int, float)):
                   count += 1
                   if showGrades:
                       print(cell_value)
               elif cell_value.strip().lower() == semester_wait.strip().lower():
                   break


   return(f"count for '{lesson}' in period '{period}': {count}")



def search_for_lesson(lesson):
    # search only in the first row (row 1)
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=1, column=col).value
        if str(cell_value).strip().lower() == lesson.strip().lower():
            return col
    return None

def search_for_starting_row(period, grade_column):
    if period.strip().lower()[0] == "m":
        for row in range(3, sheet.max_row + 1):  # loop down the column
            cell_value = sheet.cell(row=row, column=grade_column).value
            if str(cell_value).strip().lower() == period.strip().lower():
                return row
    return None
        
    
averages_total = 0
averages_count = 0

# do da thing
if command_main.strip().lower() == "avg":
    if not command_lesson.strip().lower() == "averages":
        print("")
        print(FIND_LESSON_AVG(command_lesson, command_period, command_showGrades))
    elif command_lesson.strip().lower() == "averages":
        for col in range(1, sheet.max_column + 1):
            lesson_name = sheet.cell(row=1, column=col).value
            if lesson_name is not None:
                if not lesson_name == "x":
                    # print(lesson_name)
                    # print(FIND_LESSON_AVG_AS_FLOAT(lesson_name, command_period, command_showGrades))
                    if not FIND_LESSON_AVG_AS_FLOAT(lesson_name, command_period, command_showGrades) == 0:
                        averages_total += FIND_LESSON_AVG_AS_FLOAT(lesson_name, command_period, command_showGrades)
                        averages_count += 1
                        if command_showGrades:    
                            print(lesson_name)
                            print(FIND_LESSON_AVG_AS_FLOAT(lesson_name, command_period, command_showGrades))
                            print("")
                else:
                    if averages_total > 0:
                        print("")
                        print(f"average for '{command_lesson}' in period '{command_period}': {averages_total / averages_count:.2f}")
                        break
                    else:
                        print("")
                        print(f"no {command_lesson} for period '{command_period}'.")
                        break
elif command_main.strip().lower() == "count":
    if not command_lesson.strip().lower() == "averages":
        print("")
        print(FIND_LESSON_COUNT(command_lesson, command_period, command_showGrades))
    elif command_lesson.strip().lower() == "averages":
        print("")
        print("command 'count' does not support 'averages' as lesson.")
else:
    print("")
    print(f"command '{command_main}' not found.")