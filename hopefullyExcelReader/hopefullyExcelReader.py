import openpyxl

workbook = openpyxl.load_workbook(r"D:\clasa9aNOTE.xlsx")

sheet = workbook.active

command_lesson = input(str("lesson: "))
command_main = input(str("mainCommand: "))


def test(lesson, mainCommand):
    if mainCommand.strip().lower() == "avg":
        grade_column = search_for_lesson(lesson)
        if grade_column is None:
            print(f"Lesson '{lesson}' not found.")
            return

        total = 0
        count = 0

        # Start from row 2 (skip lesson name in row 1)
        for row in range(3, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=grade_column).value
            if cell_value is not None:  # stop when empty cell
                if isinstance(cell_value, (int, float)):
                    print(cell_value)
                    total += cell_value
                    count += 1
                else:
                    break

        if count > 0:
            average = total / count
            print(f"Average for {lesson}: {average}")
        else:
            print("No valid grades found.")
    
    print("Unknown command")
    return



def search_for_lesson(lesson):
    # search only in the first row (row 1)
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=1, column=col).value
        if str(cell_value).strip().lower() == lesson.strip().lower():
            return col
    return None

# do da thing
test(command_lesson, command_main)